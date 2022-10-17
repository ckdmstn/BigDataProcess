#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook( "student.xlsx" )
ws = wb['Sheet1']

row_id = 1;
number = 0;
score = {} 
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		if sum_v not in score:
			score[sum_v] = 1
		else:
			score[sum_v] += 1
		number += 1
	row_id += 1

sorted_score = sorted(score.items(), key = lambda item: item[0], reverse = True)
dictionary = dict(sorted_score)
score_list = list(dictionary.keys())
grade = [int(number*0.15), int(number*0.3), int(number*0.5), int(number*0.7), int(number*0.85)]

count = 0
grade_dic = { 'A' : 0, 'B' : 0, 'C' : 0 }
for i in range(len(score_list)):
	row_id = 1
	score = float(score_list[i])
	count += int(dictionary[score_list[i]])
	stGrade = ""
	if count <= grade[1]:
		if count <= grade[0]:
			stGrade = "A+"
		else:
			stGrade = "A0"
		grade_dic['A'] += int(dictionary[score_list[i]])
	elif count <= grade[3]:
		n = grade[3] - grade_dic['A']
		if int(dictionary[score_list[i]]) <= n / 2 - grade_dic['B']:
			stGrade = "B+"
		else:
			stGrade = "B0"
		grade_dic['B'] += int(dictionary[score_list[i]])
	else:
		n = number - grade_dic['A'] - grade_dic['B']
		if int(dictionary[score_list[i]]) <= n / 2 - grade_dic['C']:
			stGrade = "C+"
		else:
			stGrade = "C0"
		grade_dic['C'] += int(dictionary[score_list[i]])

	for row in ws:
		if ws.cell(row = row_id, column = 7).value == score:
			ws.cell(row = row_id, column = 8).value = stGrade
		row_id += 1

wb.save( "student.xlsx" )
